import os
import sys
import random
import chardet
from copy import deepcopy
from ui.ui_main import Ui_MainWindow
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
import openpyxl

from mylogclass import MyLogClass


class MainWindow(QMainWindow, Ui_MainWindow):

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.retranslateUi(self)
        self.setFixedSize(self.width(), self.height())

        self.log = MyLogClass()

        self.signals()

    def signals(self):
        self.pushButton_excel.clicked.connect(lambda: self.open_file(self.lineEdit_excel, 'Excel (*.xlsx)'))
        self.pushButton_txt.clicked.connect(lambda: self.open_file(self.lineEdit_txt, '文本文件 (*.txt)'))
        self.pushButton_run.clicked.connect(self.run)

    # 槽函数-----------------------------------------------------------------------------------------------------------
    # 打开文件
    def open_file(self, lineEdit, file_type):
        try:
            filename, _ = QFileDialog.getOpenFileName(self, '选取文件', './', file_type)
            lineEdit.setText(filename)
        except Exception as e:
            self.log.logger.warning(str(e))

    # 生成主逻辑
    def run(self):
        try:
            path_excel = self.lineEdit_excel.text()
            path_txt = self.lineEdit_txt.text()
            if not all([path_excel, path_txt]): return

            txt_list = self.read(path_txt)

            self.wb = openpyxl.load_workbook(path_excel)
            self.ws = self.wb.active

            row_count = self.ws.max_row
            column_count = self.ws.max_column

            for r in range(1, row_count+1):
                for c in range(1, column_count+1):
                    if self.ws.cell(r, c).value and self.ws.cell(r,c).value == 'EAN':
                        self.ws.cell(r,c-1,txt_list.pop())

            self.wb.save(path_excel)
            self.write(path_txt,txt_list)

            QMessageBox.information(self, '提示', '运行完成')
        except Exception as e:
            QMessageBox.warning(self, '错误', str(e))
            self.log.logger.warning(str(e))



    # 检测编码格式，读取内容
    def read(self, path):
        try:
            with open(path, 'rb') as f:
                text = f.read()
            encode = chardet.detect(text).get('encoding')
            return [li.strip() for li in text.decode(encode).split('\n') if li.strip()]
        except Exception as e:
            raise ValueError(str(e))

    def write(self, path, data: list):
        try:
            with open(path, 'w', encoding='utf8') as f:
                f.write('\n'.join(data))
        except Exception as e:
            raise ValueError(str(e))

    def save_excel(self, filename, data_new):
        try:
            os.makedirs('./Excel/', exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            [ws.append(li) for li in data_new]
            wb.save('./Excel/' + filename + '.xlsx')
        except Exception as e:
            raise ValueError(str(e))

    # 生成excel
    def create_excel(self, img_dir, img_path):
        try:
            uid = self.create_random(self.prefix)
            data_new = deepcopy(self.data_old)
            data_new[3][1] = 'SCHVOL_{id}_%s-P' % uid
            data_new[4][1] = 'SCHVOL_{id}_%s' % uid
            d = self.data_d.pop(random.randint(0, len(self.data_d) - 1))
            f = self.data_f.pop(random.randint(0, len(self.data_f) - 1))
            data_new[3][3] = d
            data_new[4][3] = d
            data_new[4][5] = f
            data_new[4][21] = 'SCHVOL_{id}_%s-P' % uid
            imgpath_list = self.read(img_path)[:-1]
            if len(imgpath_list) > 9:
                for i in range(len(imgpath_list) - 1, 8, -1):
                    data_new[0].insert(19, '')
                    data_new[1].insert(19, 'Other Image URL{}'.format(i))
                    data_new[2].insert(19, 'other_image_url{}'.format(i))
                    data_new[3].insert(19, '')
                    data_new[4].insert(19, '')
            for n, imgpath in enumerate(imgpath_list):
                data_new[3][10 + n] = os.path.join(self.domain, imgpath).replace('\\', '/')
                data_new[4][10 + n] = os.path.join(self.domain, imgpath).replace('\\', '/')
            self.save_excel(img_dir, data_new)
        except Exception as e:
            raise ValueError(str(e))


if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
